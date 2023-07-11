from re import compile, IGNORECASE
from io import BytesIO
import time
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from .countries import get_country_iso2

platforms = {
    "Indoor": (2, 4),
    "Outdoor and Industrial": (2, 3),
}
model_pat = compile(r"^(AIR-|IW|C[9W]).*", IGNORECASE)
MAX_COL = 160
_cn_to_iso2_cache = {}


def get_book(xls_file: BytesIO) -> Workbook:
    book = load_workbook(xls_file, read_only=True, data_only=True)
    return book


def get_platform_sheet(wb: Workbook, platform: str) -> Worksheet:
    try:
        return wb[platform]
    except Exception as sheet_error:
        if "&" not in platform:
            raise sheet_error
        return wb[platform.replace("&", "and")]


def parse_row(raw_row: tuple[str], cols: dict[str, int]) -> dict[str, str]:
    vals = {}
    for header, col in cols.items():
        vals[header] = raw_row[col-1]
    return vals


def platform_indexes(platform_sheet: Worksheet) -> tuple[int, int]:
    global platforms
    if indexes := platforms.get(platform_sheet.title, None):
        rd_col = indexes[1]
        cn_col = indexes[0]
        return cn_col, rd_col
    else:
        raise Exception(f"Unable to locate sheet info for {platform_sheet.title}?")


def parse_models(platform_sheet: Worksheet, start_col: int) -> dict[str, list[int]]:
    models = {}
    try:
        for col in range(start_col, MAX_COL):
            txt = platform_sheet.cell(1, col).value
            if model_pat.search(txt):
                models[txt.upper()] = col
            else:
                break
    except IndexError:
        pass
    return models


def validate_headers(platform_sheet: Worksheet, cn_col: int, rd_col: int):
    h = {"Country": cn_col, "Regulatory Domain": rd_col}
    for header, col in h.items():
        if val := platform_sheet.cell(1, col).value:
            if val == header:
                continue  # all in order
        msg = f"Found unexpected header at 1,{col}: {val} vs {header}"
        raise ValueError(msg)


def parse_platform_models(wb: Workbook, platform: str) -> list[dict[str, object]]:
    global _cn_to_iso2_cache
    sheet = get_platform_sheet(wb, platform)
    cn_col, rd_col = platform_indexes(sheet)
    models = parse_models(sheet, rd_col+1)
    cols = models.copy()
    cols["CN"] = cn_col
    cols["RD"] = rd_col
    platform_models = []
    validate_headers(sheet, cn_col, rd_col)
    for raw_row in sheet.iter_rows(3+1, sheet.max_row+1, values_only=True):
        t1 = time.time()
        row = parse_row(raw_row, cols)
        cn = row.pop("CN")
        rd = row.pop("RD")
        if not cn or not rd:
            continue
        t2 = time.time()
        if cn_iso2 := _cn_to_iso2_cache.get(cn.lower(), None):
            pass
        else:
            cn_iso2 = get_country_iso2(cn)
            _cn_to_iso2_cache[cn.lower()] = cn_iso2
        t3 = time.time()
        pm_base = {"CN": cn_iso2, "RD": rd}
        for model, value in row.items():
            if value == "x":
                pm = pm_base.copy()
                pm["model"] = model
                platform_models.append(pm)
        # print(f"{cn}/{rd} in {time.time()-t1:.2f} s - row: {t2-t1:.2f}s, cn({cn}->{cn_iso2}): {t3-t2:.2f}s, row_scan: {time.time()-t3:.2f}s..")
    return platform_models


__all__ = ["parse_platform_models", "get_book", "platforms"]
